# 엑셀에 넣었던 데이터들을 다시 가져와서 출력 연습
from openpyxl import load_workbook
read_workbook = load_workbook("./cheong.xlsx")
read_cell = read_workbook.active
result_list = []

for i in range(1, 91):
    print(read_cell.cell(i,1).value)
    result_list.append(read_cell.cell(i,1).value)

print(result_list)

for i in result_list:
    print(i)