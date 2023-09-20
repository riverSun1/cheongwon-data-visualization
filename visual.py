# 데이터 시각화
from wordcloud import WordCloud
from openpyxl import load_workbook
from konlpy.tag import Kkma
import collections
import matplotlib.pyplot as plt
from wordcloud import WordCloud
from matplotlib import font_manager as fm
import numpy as np
from os import path
from PIL import Image
import os

Kkma = Kkma()
list_temp = []

read_wb = load_workbook("./cheong.xlsx")
read_ws = read_wb.active

list_excel = []

for i in range(1, 91):
    # print(read_ws.cell(i,1).value)
    list_excel.append(read_ws.cell(i,1).value)
    
# print(list_excel)

for row in list_excel:
    # print(Kkma.nouns(row))
    list_temp = list_temp + Kkma.nouns(row)

# print(list_temp)

result_list = []

# 안에있는 것들을 한개씩 뽑아온다.
for check in list_temp:
    # print(check)
    # 잘린 한글자 문자들 제거
    if len(check) > 1:
        result_list.append(check)

# print(result_list)

list_data = collections.Counter(result_list).most_common(10)

list_string = []
list_number = []

for s, n in list_data:
    #print(s, n)
    list_string.append(s)
    list_number.append(n)

# print(list_string)
# print(list_number)

last_text = ""

for i in list_string:
    last_text = last_text + " " + i

print(last_text)

# wordcloud = WordCloud(font_path='c:/Windows/Fonts/malgun.ttf', background_color='black', colormap='autumn', width=800, height=800)
wordcloud = WordCloud(font_path='c:/Windows/Fonts/malgun.ttf', background_color='black', colormap='prism')

wordcloud.generate(last_text)

# words_ : 객체의 비율의 정보가 담긴 딕셔너리를 반환
# wordcloud.words_

plt.figure(figsize=(10, 10))
plt.imshow(wordcloud, interpolation='bilinear')
plt.tight_layout(pad=0)
plt.axis('off')
plt.show()

mask = Image.new("RGBA",(1200,1200), (255,255,255)) # (2555,2575)는 사진 크기, (255,255,255)는 색을의미
image = Image.open('./heart.png').convert("RGBA")
x,y = image.size
mask.paste(image,(0,0,x,y),image)
mask = np.array(mask) # 픽셀 값 배열 형태 변환

wordcloud = WordCloud(font_path='c:/Windows/Fonts/malgun.ttf', mask = mask, background_color='white', colormap='prism', prefer_horizontal = True, width=800, height=800).generate(last_text)

plt.figure(figsize=(10, 10))
plt.imshow(wordcloud, interpolation='bilinear')
plt.axis('off')
plt.show()