from openpyxl import load_workbook
from konlpy.tag import Kkma
import collections

Kkma = Kkma()
list_temp = []

read_wb = load_workbook("./cheong.xlsx")
read_ws = read_wb.active

list_excel = []

for i in range(1, 91):
    # print(read_ws.cell(i,1).value)
    list_excel.append(read_ws.cell(i,1).value)
    
# print(list_excel)

for row in list_excel:
    # print(Kkma.nouns(row))
    list_temp = list_temp + Kkma.nouns(row)

# print(list_temp)

result_list = []

# 안에있는 것들을 한개씩 뽑아온다.
for check in list_temp:
    # print(check)
    # 잘린 한글자 문자들 제거
    if len(check) > 1:
        result_list.append(check)

# print(result_list)

list_data = collections.Counter(result_list).most_common(10)

list_string = []
list_number = []

for s, n in list_data:
    #print(s, n)
    list_string.append(s)
    list_number.append(n)

print(list_string)
print(list_number)